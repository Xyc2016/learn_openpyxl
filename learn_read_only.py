from openpyxl.worksheet._read_only import ReadOnlyWorksheet
import xlrd
from openpyxl import load_workbook


def load_by_openpyxl():
    a = load_workbook("a.xlsx", read_only=True)
    ws : ReadOnlyWorksheet = a.active
    print(ws.max_column, ws.max_row)
    for row_index, row in enumerate(ws.values, start=1):
        # print(row_index)
        # print(row)
        pass

def load_by_xlrd():
    a = xlrd.open_workbook("a.xlsx")
    ws = a.sheet_by_index(0)
    print(ws.ncols, ws.nrows)

load_by_xlrd()

load_by_openpyxl()