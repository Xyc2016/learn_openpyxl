import io
import xlrd
from openpyxl import load_workbook
from learn_mem import print_mem
# import xlrd

filename = "out/test_file_size.xlsx"


def open_with_openpyxl():
    wb = load_workbook(filename)
    ws = wb.active
    print(ws.max_row, ws.max_column)
    # for row_index, row in enumerate(ws.values, start=1):
        # print(row_index, row)


def open_with_xlrd():
    ws = xlrd.open_workbook(filename).sheet_by_index(0)
    nrows = ws.nrows
    ncols = ws.ncols
    print(ws.nrows, ws.ncols)


# import os
# os.system(f"open {filename}")
print_mem("before")
open_with_openpyxl()
print_mem("after")

print("?")
