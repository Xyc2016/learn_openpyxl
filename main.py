import resource
import tracemalloc
import linecache
import gc
from random import random, randint
from faker import Faker
import time
from datetime import datetime
import psutil
import io
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl.worksheet._write_only import WriteOnlyWorksheet, WriteOnlyCell
from openpyxl.styles import Font
from zipfile import ZipFile
import os
print(os.getpid())

write_only = True
wb = Workbook(
        write_only=write_only
    )
if write_only:
    ws = wb.create_sheet()
else:
    ws = wb.active

print(wb.write_only)
faker = Faker()


def print_peak_rss(prefix: str = ""):
    print(f"{prefix}:resource= {resource.getrusage(resource.RUSAGE_SELF).ru_maxrss/1024/1024}")


# ... run your application ...


process = psutil.Process(os.getpid())


def print_rss(prefix: str = ""):
    print(prefix, process.memory_info().rss / 1024 / 1024, "MB")


def random_chinese():
    val = randint(0x4e00, 0x9fbf)
    return chr(val)


# ws.column_dimensions["B"].width = 60

print_rss()
print_peak_rss()
singleton_s = "abcdefghijklmnopqrstuvwxyz"
l1 = [singleton_s] * 10
l = [1,2,3,4,5,6,7]
for row_index in range(3000):

    values = []
    values.extend(l1)
    values.extend(l)
    ws.append(values)
    # time.sleep(0.01)
    # print_rss("in progress current")
    # print_peak_rss("in progress")

print_peak_rss()
file = io.BytesIO()
gc.collect()
print_rss("before save")
wb.save(file)
print_rss("after save")
print_peak_rss("after all")

file.seek(0)
with open("a.xlsx", "wb") as file1:
    file1.write(file.read())


from openpyxl import load_workbook

wb = load_workbook("", read_only=True)
wb.active.cell(1,1)